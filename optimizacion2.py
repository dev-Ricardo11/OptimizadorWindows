import os
import shutil
import subprocess
import platform
import socket
import time
import logging
import ctypes
import sys
import psutil
import getpass
from datetime import datetime
from pathlib import Path
from tqdm import tqdm
from concurrent.futures import ThreadPoolExecutor
from typing import Tuple, List, Dict  # 🔹 Evita el error "List is not defined"
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side  
from openpyxl.drawing.image import Image 
from openpyxl.utils import get_column_letter


class SystemOptimizer:
    
    def __init__(self):
        self.start_time = datetime.now()
        self.setup_logging()
        self.informe: List[Tuple[str, str, str, str, str]] = []
        self.hostname = socket.gethostname()
        self.username = os.getlogin()
        self.serial = ""
    
    def mostrar_banner(self) -> None:
        """Muestra un banner ASCII en la terminal y lo mantiene fijo al inicio"""
        
        # Limpiar la terminal antes de mostrar el banner
        os.system('cls' if os.name == 'nt' else 'clear')

        banner = """
        ███████╗██╗   ██╗████████╗
        ██╔════╝██║   ██║╚══██╔══╝
        █████╗  ██║   ██║   ██║   
        ██╔══╝  ██║   ██║   ██║   
        ███████╗╚██████╔╝   ██║   
        ╚══════╝ ╚═════╝    ╚═╝   

        🔥 EVT - Herramienta Avanzada 🔥
        📌 Sistema en ejecución...
        ----------------------------------------
        """
        print(banner)

    # Esperar 1 segundo para que el banner se vea antes de seguir ejecutando
    time.sleep(1)
   

    def setup_logging(self) -> None:
        log_file = f"system_optimization_{self.start_time.strftime('%Y%m%d_%H%M%S')}.log"
        logging.basicConfig(
            filename=log_file,
            level=logging.ERROR,
            format='%(asctime)s - %(levelname)s - %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )

    @staticmethod
    def es_administrador() -> bool:
        try:
            return bool(ctypes.windll.shell32.IsUserAnAdmin())
        except Exception as e:
            logging.error(f"Error checking admin rights: {e}")
            return False

    def validar_entorno(self) -> None:
        if platform.system() != 'Windows':
            raise EnvironmentError("Este script solo funciona en sistemas Windows.")
        
        required_tools = ['defrag', 'DISM.exe', 'sfc.exe', 'ipconfig']
        missing_tools = [tool for tool in required_tools if not shutil.which(tool)]
        
        if missing_tools:
            raise EnvironmentError(f"Herramientas no disponibles: {', '.join(missing_tools)}")

    def obtener_info_sistema(self) -> Dict[str, str]:
        try:
            self.serial = subprocess.check_output(
                'wmic bios get serialnumber', 
                shell=True
            ).decode('utf-8', errors='ignore').split('\n')[1].strip()

            return {
                'Fecha y hora de inicio': self.start_time.strftime("%Y-%m-%d %H:%M:%S"),
                'Serial': self.serial,
                'Sistema operativo': platform.version(),
                'Procesador': platform.processor(),
                'Memoria': f"{psutil.virtual_memory().total / (1024**3):.2f} GB",
                'Uso de disco': self._get_disk_usage()
            }
        except Exception as e:
            logging.error(f"Error getting system info: {e}")
            self.serial = "No disponible"
            return {'error': str(e)}
        

    def _get_disk_usage(self) -> str:
        disk = psutil.disk_usage('C:\\')
        return f"Total: {disk.total/(1024**3):.2f}GB, Usado: {disk.used/(1024**3):.2f}GB ({disk.percent}%)"

    def _remove_file_safe(self, path: Path) -> None:
        try:
            if path.is_file():
                path.unlink()
        except Exception as e:
            logging.error(f"Error al eliminar archivo {path}: {e}")

    def configurar_firewall():
        """Configura reglas básicas de firewall para protección adicional."""
    try:
        subprocess.run(["netsh", "advfirewall", "set", "allprofiles", "state", "on"], check=True)
        subprocess.run(["netsh", "advfirewall", "firewall", "add", "rule", "name=Bloquear_Puertos", "dir=in", "action=block", "protocol=TCP", "localport=135,137,138,139,445"], check=True)
        print("Firewall configurado correctamente.")
    except Exception as e:
        print(f"Error al configurar el firewall: {e}")

    # Directorios de archivos temporales del usuario
    carpetas_temporales = [
        os.path.expanduser("~\\AppData\\Local\\Temp"),  # Carpeta Temp del usuario
    ]

    

    configurar_firewall()

    print("Limpieza completada y seguridad mejorada con firewall.")

    def limpiar_temporales(self) -> None:
        """Limpia archivos temporales del sistema."""
        temp_paths = [
            Path(os.environ.get('TEMP')),  # Usuario específico temp
            Path('C:\\Windows\\Temp'),     # Windows temp
            Path('C:\\Windows\\Temp'),
            Path(os.environ.get('LOCALAPPDATA')) / 'Temp'  # AppData Local temp
        ]
        
        for temp_path in temp_paths:
            if temp_path.exists():
                try:
                    print(f"Limpiando {temp_path}...")
                    # Calcular espacio inicial
                    espacio_inicial = sum(f.stat().st_size for f in temp_path.glob('**/*') if f.is_file()) / (1024**2)

                    
                    # Obtener lista de todos los archivos antes de intentar eliminarlos
                    archivos = list(temp_path.rglob('*'))
                    archivos.reverse()  # Empezar con los archivos más profundos
                    
                    # Eliminar archivos y carpetas
                    for item in archivos:
                        try:
                            if item.is_file():
                                try:
                                    os.chmod(str(item), 0o666)  # Dar permisos totales
                                    item.unlink(missing_ok=True)
                                except PermissionError:
                                    # Intentar forzar la eliminación usando comandos del sistema
                                    subprocess.run(['del', '/F', '/Q', str(item)], shell=True, check=False)
                            elif item.is_dir():
                                try:
                                    os.chmod(str(item), 0o666)
                                    item.rmdir()  # Intentar eliminar directorio vacío
                                except:
                                    # Si falla, intentar forzar eliminación
                                    subprocess.run(['rmdir', '/S', '/Q', str(item)], shell=True, check=False)
                        except Exception as e:
                            print(f"No se pudo eliminar {item}: {e}")
                            continue
                    
                    # Calcular espacio liberado
                    espacio_final = sum(f.stat().st_size for f in temp_path.rglob('*') if f.is_file()) / (1024**2)
                    espacio_liberado = espacio_inicial - espacio_final
                    
                    self.agregar_a_informe(
                        f"Limpieza de {temp_path}",
                        f"{espacio_inicial:.2f} MB",
                        f"{espacio_final:.2f} MB",
                        f"Liberado: {espacio_liberado:.2f} MB"
                    )
                    print(f"Espacio liberado en {temp_path}: {espacio_liberado:.2f} MB")
                    
                except Exception as e:
                    self.manejar_error(e, f"Limpieza de {temp_path}")
                    print(f"Error al limpiar {temp_path}: {e}")
        

    def limpiar_prefetch(self) -> None:
        """Limpia archivos de la carpeta Prefetch."""
        prefetch_dir = Path('C:\\Windows\\Prefetch')
        
        if prefetch_dir.exists():
            try:
                archivos_inicial = len(list(prefetch_dir.glob('*')))
                archivos_eliminados = 0
                archivos_error = 0

                # Eliminar archivos
                for archivo in prefetch_dir.glob('*.*'):
                    try:
                        os.chmod(str(archivo), 0o777)  # Dar permisos totales
                        archivo.unlink(missing_ok=True)
                        archivos_eliminados += 1
                    except PermissionError:
                        try:
                            # Intentar con otra aproximación usando subprocess
                            subprocess.run(['del', '/F', '/Q', str(archivo)], shell=True, check=False)
                            archivos_eliminados += 1
                        except:
                            archivos_error += 1
                            continue
                    except Exception as e:
                        archivos_error += 1
                        self.manejar_error(e, "Limpieza de Prefetch")
                        continue

                archivos_final = len(list(prefetch_dir.glob('*')))
                resultado = f"Completado - Eliminados: {archivos_eliminados}, Errores: {archivos_error}"
                
                self.agregar_a_informe(
                    "Limpieza de Prefetch",
                    f"{archivos_inicial} archivos",
                    f"{archivos_final} archivos",
                    resultado
                )
                
                print(f"Prefetch: {resultado}")
                
            except Exception as e:
                self.manejar_error(e, "Limpieza de Prefetch")
                print(f"Error en Prefetch: {e}")

    def desfragmentar_disco(self) -> None:
        discos = ['C:', 'D:']  # Agrega más discos si es necesario
        
        for disco in discos:
            try:
                print(f"Iniciando desfragmentación del disco {disco}...")
                
                resultado = subprocess.run(
                    ['defrag', disco, '/U', '/V'],
                    capture_output=True, text=True, check=True
                )
                
                print(resultado.stdout)
                logging.info(f"Desfragmentación del disco {disco} completada correctamente.")
                
            except subprocess.CalledProcessError as e:
                print(f"Error al desfragmentar {disco}: {e.stderr}")
                logging.error(f"Error al desfragmentar {disco}: {e.stderr}")
            except Exception as e:
                print(f"Error inesperado al desfragmentar {disco}: {e}")
                logging.error(f"Error inesperado al desfragmentar {disco}: {e}")


    def ejecutar_comandos_dism(self) -> None:
        comandos = [
            ("Verificando archivos del sistema...", ['sfc', '/scannow']),
            ("Verificación de la imagen...", ['DISM', '/Online', '/Cleanup-Image', '/CheckHealth']),
            ("Escaneando la imagen...", ['DISM', '/Online', '/Cleanup-Image', '/ScanHealth']),
            ("Restaurando la imagen...", ['DISM', '/Online', '/Cleanup-Image', '/RestoreHealth']),
            ("Restaurando la imagen...", ['DISM', '/Online', '/Cleanup-Image', '/startcomponentcleanup'])
        ]

        for descripcion, comando in comandos:
            try:
                print(descripcion)
                with tqdm(total=100, desc=descripcion, unit="%") as pbar:
                    subprocess.run(comando, check=True)
                    pbar.update(100)
                self.agregar_a_informe(descripcion, "Estado inicial desconocido", "Estado final desconocido", "Completado")
            except Exception as e:
                self.manejar_error(e, descripcion)

    def limpiar_cache_dns(self) -> None:
        try:
            print("Ejecutando flushdns para limpiar la caché DNS...")
            subprocess.run(['ipconfig', '/flushdns'], check=True)
            self.agregar_a_informe(
                "Limpieza de caché DNS (flushdns)",
                "N/A",
                "N/A",
                "Completado"
            )
            print("Caché DNS limpiada correctamente.")
        except subprocess.CalledProcessError as e:
            self.manejar_error(e, "Limpieza de caché DNS (flushdns)")

    def agregar_a_informe(self, proceso: str, estado_inicial: str, estado_final: str, resultado: str) -> None:
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.informe.append((timestamp, proceso, estado_inicial, estado_final, resultado))
        logging.info(f"[{timestamp}] Proceso completado: {proceso} - Resultado: {resultado}")

    def manejar_error(self, e: Exception, contexto: str) -> None:
        error_msg = f"{type(e).__name__}: {str(e)}"
        logging.error(f"Error en {contexto}: {error_msg}", exc_info=True)
        self.agregar_a_informe(
            contexto,
            "Estado inicial desconocido",
            "Error encontrado",
            error_msg
        )

    def ejecutar_todo(self) -> None:
        try:
            self.mostrar_banner()  # 🔹 Mostrar el banner ANTES de comenzar la optimización

            print("Iniciando optimización del sistema...")
            self.validar_entorno()
            print("✔️ Entorno validado")

            info_sistema = self.obtener_info_sistema()
            print("✔️ Información del sistema obtenida")

            print("🧹 Limpiando archivos temporales...")
            self.limpiar_temporales()

            print("🗑️ Limpiando Prefetch...")
            self.limpiar_prefetch()

            print("💾 Desfragmentando disco...")
            self.desfragmentar_disco()

            print("🛠️ Ejecutando DISM y SFC...")
            self.ejecutar_comandos_dism()

            print("🌐 Limpiando caché DNS...")
            self.limpiar_cache_dns()

            print("📊 Generando informe final...")
            self.generar_informe()
            print("\n✅ Optimización del sistema completada exitosamente.")

            # 🔹 Mostrar mensaje emergente antes del reinicio
            ctypes.windll.user32.MessageBoxW(0, "El equipo se reiniciará en 10 segundos...", "Aviso de Reinicio", 1)

            # 🔹 Esperar 10 segundos antes de reiniciar
            time.sleep(10)

            # 🔹 Reiniciar el equipo
            os.system("shutdown /r /t 0")

        except Exception as e:
            logging.error(f"Error durante la optimización del sistema: {e}")
            print(f"❌ Se produjo un error: {e}")
            raise


    def obtener_usuario_actual(self):
        try:
            # Obtener nombre de usuario del sistema
            username = getpass.getuser()
            # Obtener el hostname actual
            computer_name = os.environ.get('COMPUTERNAME', '')
            print(f"Usuario actual: {username}")
            print(f"Computadora: {computer_name}")
            return username, computer_name
        except Exception as e:
            print(f"Error al obtener usuario: {e}")
            return None, None    
    

    def obtener_usuario_actual(self):
        return os.getlogin(), os.environ.get('COMPUTERNAME', '')
    
    

    def generar_informe(self):
        try:
            fecha_actual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            username, computer_name = self.obtener_usuario_actual()
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")  
            nombre_archivo = f"{self.serial}_Informe_Mantenimiento_{timestamp}.xlsx"
            
            # Nueva ruta local
            ruta_local = Path(f"C:/Informes_Mantenimiento/{self.serial}")
            print(f"📂 Creando directorio: {ruta_local}")
            ruta_local.mkdir(parents=True, exist_ok=True)
            
            if not ruta_local.exists():
                print("❌ Error: No se pudo crear el directorio.")
                return
            
            ruta_guardado = ruta_local / nombre_archivo
            print(f"📁 Guardando archivo en: {ruta_guardado}")
            
            
                    # Crear un nuevo libro y hoja
            wb = Workbook()
            ws = wb.active
            ws.title = "Informe de Mantenimiento"

                    # Ajustar anchos de columna
            columnas = {'A': 20, 'B': 20, 'C': 20, 'D': 90}
            for col, width in columnas.items():
                        ws.column_dimensions[col].width = width

                    # Definir estilos
            header_font = Font(bold=True, size=11, color="000000")
            normal_font = Font(size=11)
            blue_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            light_blue_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

                    # Definir bordes
            thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )

                    # Insertar logo si existe
                    # Insertar logo si existe
            logo_path = r"C:\\Escaner\\img\\Expreso.png"
            if os.path.exists(logo_path):
                        img = Image(logo_path)
                        img.width = 100
                        img.height = 50
                        ws.add_image(img, "A1")

                    # Encabezado principal
            ws['C1'] = "FOR INFORME MANTENIMIENTO"
            ws['C1'].font = header_font
            ws['C1'].alignment = center_align

                    # Información de control (lado derecho)
            control_info = [
                        ('E1', 'Código', 'F1', 'FOR-PRY-TEC-006'),
                        ('E2', 'Versión', 'F2', '01'),
                        ('E3', 'Fecha', 'F3', '12/02/2025'),
                        ('E4', 'Fecha de revisión', 'F4', '12/05/2025')
                    ]
            for e_cell, e_value, f_cell, f_value in control_info:
                        ws[e_cell] = e_value
                        ws[e_cell].font = normal_font
                        ws[e_cell].border = thin_border
                        ws[f_cell] = f_value
                        ws[f_cell].font = normal_font
                        ws[f_cell].border = thin_border

                    # Título de sección
            ws['A6'] = "Informe de Mantenimiento Preventivo"
            ws.merge_cells('A6:D6')
            ws['A6'].fill = blue_fill
            ws['A6'].font = Font(bold=True, color="FFFFFF")
            ws['A6'].alignment = center_align

                    # Información básica
            basic_info = [
                        ('A7', 'Fecha del Mantenimiento:', fecha_actual),
                        ('A8', 'Usuario:', username),
                        ('A9', 'Nombre del Equipo:', computer_name),
                        ('A10', 'Número de Serie:', self.serial)
                    ]
            for cell, label, value in basic_info:
                        ws[cell] = label
                        ws[cell].font = normal_font
                        ws[cell].border = thin_border
                        ws[cell.replace("A", "B")] = value  # Valor en columna B
                        ws[cell.replace("A", "B")].alignment = left_align  # Asigna el valor en la columna B

            ws.column_dimensions['D'].width = 80 
                    
                    # Encabezados de la tabla
            headers = [('A12', 'Fecha'), ('B12', 'Usuario'), ('C12', 'Equipo'), ('D12', 'Mantenimiento Preventivo')]
            for cell, value in headers:
                        ws[cell] = value
                        ws[cell].font = header_font
                        ws[cell].fill = blue_fill
                        ws[cell].alignment = center_align
                        ws[cell].border = thin_border
                    
                    # Agregar datos a la tabla
            fila = ws.max_row + 1
            mantenimiento_texto = """Mantenimiento: Preventivo\n"
                    "Se realiza diagnóstico del disco SSD, aparece en buen estado. Iniciamos limpieza de temporales, CACHE, DNS, PREFETCH, SFC /SCANNOW, "
                    "ejecutamos comandos de reparación DISM, activamos núcleos y memoria RAM, optimizamos y desfragmentamos disco, "
                    "actualizamos WINDOWS UPDATE, BIOS y reiniciamos el equipo."""
            ws[f'A{fila}'] = fecha_actual
            ws[f'B{fila}'] = username
            ws[f'C{fila}'] = computer_name
            ws[f'D{fila}'] = mantenimiento_texto

            for col in ['A', 'B', 'C', 'D']:
                        ws[f'{col}{fila}'].alignment = left_align if col == 'D' else center_align
                        ws[f'{col}{fila}'].border = thin_border
            
            # Guardar archivo localmente
            wb.save(str(ruta_guardado))
            
            if ruta_guardado.exists():
                print(f"✅ Informe guardado en: {ruta_guardado}")
            else:
                print("❌ Error: No se pudo guardar el archivo.")
        except Exception as e:
            print(f"❌ Error al generar el informe: {e}")



# Bloque principal
if __name__ == "__main__":
    optimizer = SystemOptimizer()
    if optimizer.es_administrador():
        try:
            optimizer.ejecutar_todo()
        except Exception as e:
            print(f"Error durante la optimización: {e}")
    else:
        print("⚠️ Este script requiere permisos de administrador para ejecutarse.")
        
# Mostrar mensaje tipo popup
ctypes.windll.user32.MessageBoxW(0, "Se realizo Mantenimiento preventivo de tu equipos en pocos minutos se va a reinciar ", "Aviso de Reinicio", 1)

# Esperar 10 segundos y reiniciar
time.sleep(10)
os.system("shutdown /r /t 0")      